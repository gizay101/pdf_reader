import pdfplumber
from openpyxl import Workbook
import os
import re

#pip install pdfplumber
#pip install xlrd
#pip install xlwt


pdf_folder = r"" #Folder containing pdf files to be extracted
excel_file = r"output.xlsx" # Output.xlsx file change to desired location

# the .xlsx file layout, change to desired extraction
#In this example I will be listing the document name
#the type of document, the date listed, the postcode found

#Creates an array
workbook = Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1, value='Document')
worksheet.cell(row=1, column=2, value='Type')
worksheet.cell(row=1, column=3, value='Date')
worksheet.cell(row=1, column=4, value='Postcode')
worksheet.cell(row=1, column=5, value='type')
#Loop going through all .pdf files in set location.
for root, dirs, files in os.walk(pdf_folder):
    for file in files:
        if file.endswith(".pdf"):
            pdf_file = os.path.join(root, file)

            with pdfplumber.open(pdf_file) as pdf:
                first_page = pdf.pages[0] #This essentially only checks the first page for my project
                text = first_page.extract_text()#Extracts all the readable elements in the first page
                pdf_name = os.path.basename(pdf_file)

                found_date = False
                found_postcode = False
                found_text = False
                found_type = False
                date = None
                postcode = None
                type = None

                #This following part adjust as you see fit, to your projects
                #These are some examples of finding text in pdf files for extraction
                #Its a good idea to run the inital_pdf_reader to see what the pdf text looks like before attempting this


                #The first text search looks for the following key words and labels the following doc type 1-4 and prints True to found_text
                #If nothing is found later prints false which results in an error outcome for later
                doc_type = None
                if "CONDITION REPORT" in text:
                    doc_type = "1"
                    found_text = True
                elif "CERTIFICATE" in text:
                    doc_type = "2"
                    found_text = True
                elif "ELECTRICAL" in text:
                    doc_type = "3"
                    found_text = True
                elif "REPORT" in text:
                    doc_type = "4"
                    found_text = True

               #The next search is for a date, the following algorithm to find a date listed
                date_match = re.search(r'\d{1,2}/\d{1,2}/\d{4}', text)
                if date_match:
                    date = date_match.group()
                    found_date = True

                #The next search is for a postcode which in my case was always between two n/a in the text
                #Finding the text embedded in a certain way can always be taken advantage of and utilise an faster way to search through text
                postcode_match = re.search(r'N/A (.*?) N/A', text)
                if postcode_match:
                    postcode = postcode_match.group(1)
                    found_postcode = True

                #The last search is looking for two conditions correct or incorrect in the text
                type_match = re.search(r'\b(correct|incorrect)\b', text, re.IGNORECASE)
                if type_match:
                    type = type_match.group(0)
                    found_type = True

                #If any searches above do not apply true and stay false, "ERROR" will be applied in the .xlsx file
                worksheet.append([pdf_name, doc_type if found_text else "ERROR",
                                  date if found_date else "ERROR",
                                  postcode if found_postcode else "ERROR",
                                  type if found_type else "ERROR"])

#Saves the array to the file.
workbook.save(excel_file)
